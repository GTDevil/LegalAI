"""Lead rows for the on-screen sheet, plus CSV/Excel import and export."""

from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

COLUMNS = [
    "name",
    "phone",
    "call_status",
    "interested",
    "total_loan_amount",
    "remaining_amount",
    "settlement_amount",
    "legal_fee",
    "fee_percent",
    "cibil_or_experience",
    "notes",
]

HEADERS = [
    "Name",
    "Phone",
    "Call status",
    "Interested",
    "Total loan amount",
    "Remaining loan amount",
    "Settlement amount offered",
    "Legal fee",
    "Fee %",
    "CIBIL / experience score",
    "Notes",
]

STATUS_NOT_CALLED = "Not called"
STATUS_CALLING = "Calling"
STATUS_NO_ANSWER = "No answer"
STATUS_NO_LOAN = "No loan"
STATUS_NOT_INTERESTED = "Not interested"
STATUS_INTERESTED = "Interested"
STATUS_COMPLETED = "Completed"
STATUS_FAILED = "Call failed"


@dataclass
class Lead:
    name: str = ""
    phone: str = ""
    call_status: str = STATUS_NOT_CALLED
    interested: str = ""
    total_loan_amount: float | None = None
    remaining_amount: float | None = None
    settlement_amount: float | None = None
    legal_fee: float | None = None
    fee_percent: float | None = None
    cibil_or_experience: str = ""
    notes: str = ""
    row_id: str = field(default="")

    def display_values(self) -> dict[str, str]:
        def money(value: float | None) -> str:
            if value is None:
                return ""
            return f"{value:.2f}"

        return {
            "name": self.name,
            "phone": self.phone,
            "call_status": self.call_status,
            "interested": self.interested,
            "total_loan_amount": money(self.total_loan_amount),
            "remaining_amount": money(self.remaining_amount),
            "settlement_amount": money(self.settlement_amount),
            "legal_fee": money(self.legal_fee),
            "fee_percent": "" if self.fee_percent is None else f"{self.fee_percent:g}",
            "cibil_or_experience": self.cibil_or_experience,
            "notes": self.notes,
        }

    def to_export_row(self) -> dict[str, Any]:
        data = asdict(self)
        data.pop("row_id", None)
        return data


def _parse_amount(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("₹", "").replace("$", "")
    if not text:
        return None
    return float(text)


def lead_from_mapping(data: dict[str, Any]) -> Lead:
    lowered = {str(key).strip().lower().replace(" ", "_"): value for key, value in data.items()}

    def pick(*keys: str, default: str = "") -> str:
        for key in keys:
            if key in lowered and lowered[key] not in (None, ""):
                return str(lowered[key]).strip()
        return default

    return Lead(
        name=pick("name"),
        phone=pick("phone", "number", "mobile"),
        call_status=pick("call_status", "status", default=STATUS_NOT_CALLED),
        interested=pick("interested"),
        total_loan_amount=_parse_amount(lowered.get("total_loan_amount") or lowered.get("total_loan")),
        remaining_amount=_parse_amount(
            lowered.get("remaining_amount") or lowered.get("remaining_loan_amount")
        ),
        settlement_amount=_parse_amount(
            lowered.get("settlement_amount") or lowered.get("settlement_amount_offered")
        ),
        legal_fee=_parse_amount(lowered.get("legal_fee") or lowered.get("fee_for_legal_concerns")),
        fee_percent=_parse_amount(lowered.get("fee_percent") or lowered.get("fee_%")),
        cibil_or_experience=pick("cibil_or_experience", "cibil", "experience_score"),
        notes=pick("notes"),
    )


def empty_leads(count: int = 8) -> list[Lead]:
    return [Lead() for _ in range(count)]


def load_csv(path: Path) -> list[Lead]:
    import csv

    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return [lead_from_mapping(row) for row in reader]


def save_csv(path: Path, leads: list[Lead]) -> None:
    import csv

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS)
        writer.writeheader()
        for lead in leads:
            writer.writerow(lead.to_export_row())


def load_xlsx(path: Path) -> list[Lead]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    leads: list[Lead] = []
    for row in rows[1:]:
        mapping = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
        lead = lead_from_mapping(mapping)
        if lead.name or lead.phone:
            leads.append(lead)
    return leads


def save_xlsx(path: Path, leads: list[Lead]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Leads"
    sheet.append(HEADERS)
    for lead in leads:
        row = lead.to_export_row()
        sheet.append([row[column] for column in COLUMNS])
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 22
    workbook.save(path)


def load_workbook_file(path: Path) -> list[Lead]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx(path)
    raise ValueError("Please choose a .xlsx or .csv file")


def save_workbook_file(path: Path, leads: list[Lead]) -> None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        save_csv(path, leads)
        return
    if suffix in {".xlsx", ".xlsm", ""}:
        if suffix == "":
            path = path.with_suffix(".xlsx")
        save_xlsx(path, leads)
        return
    raise ValueError("Please save as .xlsx or .csv")
